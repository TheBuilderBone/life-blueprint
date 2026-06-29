#!/usr/bin/env python3
"""
Import SoFi checking CSV into Budget_System.xlsx Transactions sheet.

Usage:
    python utilities/import_txns.py bank.csv
    python utilities/import_txns.py bank.csv --months 3   # last N months only

SoFi CSV columns: Date, Description, Type, Amount, Current balance, Status
- Keeps: DEBIT_CARD, DIRECT_PAY
- Skips: DEPOSIT, DIRECT_DEPOSIT, WITHDRAWAL, INTEREST_EARNED, CHECK, ATM, TRANSFER
- Flips negative amounts to positive
- Deduplicates against rows already in the Transactions sheet
- Clears sample rows (those whose description matches a known sample merchant)
"""

import os
import sys
import csv
import datetime
import argparse

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

WORKBOOK = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Budget_System.xlsx")

KEEP_TYPES   = {"DEBIT_CARD", "DIRECT_PAY"}
SAMPLE_DESCS = {"SHELL #1234 JACKSONVILLE", "WALMART SUPERCENTER", "MCDONALDS F12"}

# Column indices in Transactions sheet (1-based)
COL_DATE = 2   # B
COL_DESC = 3   # C
COL_AMT  = 4   # D

DATA_START_ROW = 6   # first data row (row 5 is header)

# ── Palette (must match build_workbook.py) ─────────────────────────────────────
ABLUE = "EFF6FF"; BLUE  = "2563EB"; WHITE = "FFFFFF"; LGRAY = "F8FAFC"
TEXT  = "1E293B"; SUB   = "64748B"; AMBER = "D97706"; GREEN = "16A34A"
DARK  = "0F172A"; LAMBER= "FEF3C7"

def fill(c):
    return PatternFill("solid", fgColor=c)

def font_style(bold=False, color=TEXT, size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align_style(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v)


def read_sofi_csv(path, months_back=None):
    """Parse SoFi CSV; return list of (date_str, description, amount)."""
    cutoff = None
    if months_back:
        today = datetime.date.today()
        y = today.year - (months_back // 12)
        m = today.month - (months_back % 12)
        if m <= 0:
            m += 12
            y -= 1
        cutoff = datetime.date(y, m, 1)

    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            txn_type = row.get("Type", "").strip()
            if txn_type not in KEEP_TYPES:
                continue
            status = row.get("Status", "").strip()
            if status.lower() not in ("posted", ""):
                continue
            try:
                amount = float(row["Amount"])
            except (ValueError, KeyError):
                continue
            if amount >= 0:
                continue  # skip credits/refunds
            amount = round(-amount, 2)  # flip to positive

            date_str = row.get("Date", "").strip()
            try:
                date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                try:
                    date_obj = datetime.datetime.strptime(date_str, "%m/%d/%Y").date()
                except ValueError:
                    continue
            if cutoff and date_obj < cutoff:
                continue

            desc = row.get("Description", "").strip()
            rows.append((date_obj.strftime("%Y-%m-%d"), desc, amount))

    rows.sort(key=lambda r: r[0])
    return rows


def get_existing_rows(ws):
    """Return set of (date_str, desc, amount) tuples already in the sheet."""
    existing = set()
    for r in range(DATA_START_ROW, ws.max_row + 1):
        b = ws.cell(r, COL_DATE).value
        c = ws.cell(r, COL_DESC).value
        d = ws.cell(r, COL_AMT).value
        if not c:
            continue
        date_str = b.strftime("%Y-%m-%d") if hasattr(b, "strftime") else str(b or "")
        existing.add((date_str, str(c).strip(), float(d or 0)))
    return existing


def _cat_fx(r):
    return (
        f"=IF($C{r}=\"\",\"\",IFERROR(LOOKUP(2,"
        f"1/((ISNUMBER(SEARCH('Category Rules'!$B$6:$B$125,$C{r})))"
        f"*('Category Rules'!$B$6:$B$125<>\"\")),"
        f"'Category Rules'!$C$6:$C$125),\"Uncategorized\"))"
    )


def write_row(ws, r, date_str, desc, amt):
    bg = LGRAY if r % 2 == 0 else WHITE

    # B: Date
    ws.cell(r, COL_DATE).value         = date_str
    ws.cell(r, COL_DATE).fill          = fill(ABLUE)
    ws.cell(r, COL_DATE).font          = font_style(color=BLUE)
    ws.cell(r, COL_DATE).alignment     = align_style("center")
    ws.cell(r, COL_DATE).number_format = "YYYY-MM-DD"

    # C: Description
    ws.cell(r, COL_DESC).value     = desc
    ws.cell(r, COL_DESC).fill      = fill(ABLUE)
    ws.cell(r, COL_DESC).font      = font_style(color=BLUE)
    ws.cell(r, COL_DESC).alignment = align_style("left")

    # D: Amount
    ws.cell(r, COL_AMT).value         = amt
    ws.cell(r, COL_AMT).fill          = fill(ABLUE)
    ws.cell(r, COL_AMT).font          = font_style(color=BLUE)
    ws.cell(r, COL_AMT).alignment     = align_style("center")
    ws.cell(r, COL_AMT).number_format = '"$"#,##0.00'

    # E: AUTO CAT formula
    ws.cell(r, 5).value     = _cat_fx(r)
    ws.cell(r, 5).fill      = fill(bg)
    ws.cell(r, 5).font      = font_style(color="7C3AED")
    ws.cell(r, 5).alignment = align_style("center")

    # F: Override (blank, user editable)
    ws.cell(r, 6).fill      = fill(LAMBER)
    ws.cell(r, 6).font      = font_style(color=AMBER)
    ws.cell(r, 6).alignment = align_style("center")

    # G: Final Cat
    ws.cell(r, 7).value     = f'=IF($C{r}="","",IF($F{r}<>"",$F{r},$E{r}))'
    ws.cell(r, 7).fill      = fill(bg)
    ws.cell(r, 7).font      = font_style(color=GREEN)
    ws.cell(r, 7).alignment = align_style("center")

    # H: Month label
    ws.cell(r, 8).value     = f'=IF($B{r}="","",TEXT($B{r},"mmm yyyy"))'
    ws.cell(r, 8).fill      = fill(bg)
    ws.cell(r, 8).font      = font_style(color=SUB)
    ws.cell(r, 8).alignment = align_style("center")


def clear_row(ws, r):
    """Blank out a row completely (remove sample data)."""
    bg = LGRAY if r % 2 == 0 else WHITE
    for col in range(1, 9):
        ws.cell(r, col).value = None
        ws.cell(r, col).fill  = fill(bg)


def main():
    parser = argparse.ArgumentParser(description="Import SoFi CSV into Budget_System.xlsx")
    parser.add_argument("csv_file", help="Path to SoFi checking CSV export")
    parser.add_argument("--months", type=int, default=None,
                        help="Only import transactions from the last N months")
    args = parser.parse_args()

    if not os.path.exists(args.csv_file):
        print(f"Error: file not found: {args.csv_file}")
        sys.exit(1)

    if not os.path.exists(WORKBOOK):
        print(f"Error: workbook not found: {WORKBOOK}")
        sys.exit(1)

    print(f"Reading {args.csv_file} ...")
    new_rows = read_sofi_csv(args.csv_file, args.months)
    print(f"  Found {len(new_rows)} spending transactions (DEBIT_CARD + DIRECT_PAY)")

    print(f"Opening {WORKBOOK} ...")
    wb = load_workbook(WORKBOOK)
    if "Transactions" not in wb.sheetnames:
        print("Error: 'Transactions' sheet not found in workbook.")
        sys.exit(1)
    ws = wb["Transactions"]

    existing = get_existing_rows(ws)
    print(f"  Existing non-empty rows: {len(existing)}")

    # Clear sample rows (rows that contain known sample data)
    cleared = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        desc_cell = ws.cell(r, COL_DESC).value
        if desc_cell and str(desc_cell).strip() in SAMPLE_DESCS:
            clear_row(ws, r)
            cleared += 1
    if cleared:
        print(f"  Cleared {cleared} sample row(s)")

    # Re-read existing after clearing samples
    existing = get_existing_rows(ws)

    # Filter to new (deduplicated) rows
    to_add = []
    for date_str, desc, amt in new_rows:
        key = (date_str, desc, amt)
        if key not in existing:
            to_add.append((date_str, desc, amt))

    print(f"  New transactions to add: {len(to_add)}")
    if not to_add:
        print("Nothing to import — all transactions already present.")
        wb.save(WORKBOOK)
        return

    # Find the first empty data row
    write_r = DATA_START_ROW
    for r in range(DATA_START_ROW, ws.max_row + 2):
        if not ws.cell(r, COL_DESC).value:
            write_r = r
            break

    # Check capacity
    max_row = 1005  # must match _TXN_DATA_END in build_workbook.py
    available = max_row - write_r + 1
    if len(to_add) > available:
        print(f"  Warning: sheet has room for {available} more rows; "
              f"truncating at {available} transactions.")
        to_add = to_add[:available]

    for date_str, desc, amt in to_add:
        write_row(ws, write_r, date_str, desc, amt)
        write_r += 1

    wb.save(WORKBOOK)
    print(f"Done. Imported {len(to_add)} transactions → {WORKBOOK}")
    print(f"  Next empty row: {write_r}")


if __name__ == "__main__":
    main()
