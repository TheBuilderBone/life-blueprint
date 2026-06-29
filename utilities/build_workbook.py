#!/usr/bin/env python3
"""
Generate Budget_System.xlsx with historical data, live formulas, and embedded charts.
Run: python utilities/build_workbook.py
"""

import os
import csv
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.drawing.image import Image as XLImage

BILLS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bills.csv")
CHARTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "charts")
OUT_FILE   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Budget_System.xlsx")

UTILITIES = ["Water", "Power", "Internet", "Trash"]
FUNDED    = 630
BUFFER    = 167.63

# ── Pre-computed budget values ─────────────────────────────────────────────────
# Python equivalents of every Excel formula in The System tab so the file
# displays correct numbers in any viewer (Google Sheets, LibreOffice, preview).
_CC        = 2560;  _SIDE = 480;   _SHANE = 628
_MORTGAGE  = 837;   _HINS = 190;   _AINS  = 79;  _PTAX = 54; _PHONE = 55
_WATER     = 90;    _POWER = 453;  _INET  = 64;  _TRASH = 23
_GAS       = 430;   _GROC  = 400;  _SPEND = 500
_LEAK_HDG  = 45;    _INET_SURP = 39

INCOME_TOTAL  = _CC + _SIDE + _SHANE                          # 3668
FIXED_TOTAL   = _MORTGAGE + _HINS + _AINS + _PTAX + _PHONE   # 1215
UTIL_SUB      = _WATER + _POWER + _INET + _TRASH              # 630
FLEX_TOTAL_PY = UTIL_SUB + _GAS + _GROC                       # 1460
LEFTOVER_PY   = INCOME_TOTAL - FIXED_TOTAL - FLEX_TOTAL_PY    # 993
SHANE_CONT    = LEFTOVER_PY - _SHANE                          # 365
SAVINGS_PY    = LEFTOVER_PY - _SPEND                          # 493
PAD_TOTAL     = _LEAK_HDG + _INET_SURP                        # 84
SE_RESERVE_PY = round(_SIDE * 0.28, 2)                        # 134.40
SE_NET_PY     = _SIDE - SE_RESERVE_PY                         # 345.60
SINK_TOTAL_PY = 75 + 17 + 0                                   # 92
EST_TH_PY     = round(22 * 40 * 52 / 12 * 0.72, 2)           # 2745.60
INCOME_GAIN_PY = round(EST_TH_PY - _CC, 2)                   # 185.60
TRUE_AVAIL_PY = round(LEFTOVER_PY - SE_RESERVE_PY - SINK_TOTAL_PY, 2)  # 766.60
SHANE_FL_PY   = round(TRUE_AVAIL_PY - _SHANE, 2)             # 138.60

# ── Palette ────────────────────────────────────────────────────────────────────
DARK  = "0F172A"; TEXT  = "1E293B"; SUB   = "64748B"
BLUE  = "2563EB"; LBLUE = "DBEAFE"; ABLUE = "EFF6FF"
AMBER = "D97706"; LAMBER= "FEF3C7"
RED   = "DC2626"; LRED  = "FEE2E2"
GREEN = "16A34A"; LGREEN= "DCFCE7"
LGRAY = "F8FAFC"; MGRAY = "E2E8F0"; WHITE = "FFFFFF"

UTIL_HEX = {"Water": "0EA5E9", "Power": "F59E0B",
            "Internet": "8B5CF6", "Trash": "10B981"}

# ── Style helpers ──────────────────────────────────────────────────────────────
def fill(c): return PatternFill("solid", fgColor=c)
def font(bold=False, color=TEXT, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_border(color=MGRAY):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def bottom_border(color=DARK):
    return Border(bottom=Side(style="medium", color=color))

def cell(ws, rc, value=None, formula=None, bold=False, clr=TEXT, size=10,
         italic=False, bg=WHITE, ha="left", va="center", wrap=False,
         num=None, border=None):
    c = ws[rc]
    c.value     = formula if formula else value
    c.font      = font(bold=bold, color=clr, size=size, italic=italic)
    c.fill      = fill(bg)
    c.alignment = align(ha, va, wrap)
    if num:    c.number_format = num
    if border: c.border = border
    return c

def header_row(ws, row, cols, labels, colors, height=26):
    ws.row_dimensions[row].height = height
    for i, (lbl, clr) in enumerate(zip(labels, colors), 1):
        c = get_column_letter(i if isinstance(cols, int) else cols[i-1])
        ws[f"{c}{row}"] = lbl
        ws[f"{c}{row}"].fill      = fill(clr)
        ws[f"{c}{row}"].font      = Font(bold=True, color=WHITE, size=9, name="Calibri")
        ws[f"{c}{row}"].alignment = align("center", "center")

def section_bar(ws, row, title, col_span="A:H", bg=DARK, height=24):
    ws.merge_cells(f"A{row}:{col_span[-1]}{row}")
    ws[f"A{row}"] = title
    ws[f"A{row}"].fill      = fill(bg)
    ws[f"A{row}"].font      = Font(bold=True, color=WHITE, size=11, name="Calibri")
    ws[f"A{row}"].alignment = align("left", "center")
    ws.row_dimensions[row].height = height

def embed_image(ws, path, anchor, w, h):
    if os.path.exists(path):
        img = XLImage(path)
        img.width  = w
        img.height = h
        ws.add_image(img, anchor)

# ── Data loading ───────────────────────────────────────────────────────────────
def load_bills():
    rows = []
    with open(BILLS_FILE, newline="") as f:
        for row in csv.DictReader(f):
            row["amount"]  = float(row["amount"])
            row["exclude"] = row["exclude"].strip().lower() in ("1","true","yes","x")
            rows.append(row)
    return rows

def get_monthly(rows):
    m = {}
    for r in rows:
        d = r["date"]
        if d not in m:
            m[d] = {u: None for u in UTILITIES}
        if r["utility"] in UTILITIES:
            m[d][r["utility"]] = r["amount"]
    return dict(sorted(m.items()))

MNAME = {"01":"Jan","02":"Feb","03":"Mar","04":"Apr","05":"May","06":"Jun",
          "07":"Jul","08":"Aug","09":"Sep","10":"Oct","11":"Nov","12":"Dec"}

def mlabel(d):
    y, m = d.split("-")
    return f"{MNAME[m]} '{y[2:]}"

def future_months(last, n=18):
    y, m = int(last[:4]), int(last[5:])
    out = []
    for _ in range(n):
        m += 1
        if m > 12: m, y = 1, y+1
        out.append(f"{y:04d}-{m:02d}")
    return out


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — The System
# ══════════════════════════════════════════════════════════════════════════════
def build_system(wb):
    ws = wb.active
    ws.title = "The System"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 36

    # ── Title ─────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    ws["A1"] = "THE SYSTEM — FIXED vs FLEX"
    ws["A1"].fill      = fill(DARK)
    ws["A1"].font      = Font(bold=True, size=18, color=WHITE, name="Calibri")
    ws["A1"].alignment = align("left", "center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:D2")
    ws["A2"] = "Flex utilities run on 12 months of YOUR real data  ·  fixed bills run themselves"
    ws["A2"].fill      = fill(LGRAY)
    ws["A2"].font      = font(italic=True, size=9, color=SUB)
    ws["A2"].alignment = align("left","center")
    ws.row_dimensions[2].height = 18

    def row(r, label, val=None, fx=None, note=None, bold=False, bg=WHITE, clr=TEXT):
        ws.row_dimensions[r].height = 20
        cell(ws, f"B{r}", label, bold=bold, bg=bg, clr=clr, va="center")
        if val is not None:
            cell(ws, f"C{r}", val,  bold=bold, bg=bg, ha="right",
                 num='"$"#,##0.00')
        elif fx:
            cell(ws, f"C{r}", formula=fx, bold=bold, bg=bg, ha="right",
                 num='"$"#,##0.00')
        else:
            ws[f"C{r}"].fill = fill(bg)
        if note:
            cell(ws, f"D{r}", note, italic=True, bg=bg, clr=SUB, size=9, wrap=True)
        else:
            ws[f"D{r}"].fill = fill(bg)

    def sec(r, title, bg=DARK):
        ws.row_dimensions[r].height = 22
        ws.merge_cells(f"A{r}:D{r}")
        ws[f"A{r}"] = title
        ws[f"A{r}"].fill      = fill(bg)
        ws[f"A{r}"].font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
        ws[f"A{r}"].alignment = align("left","center")

    r = 4
    sec(r, "INCOME  (spendable)"); r+=1
    row(r, "Coca-Cola take-home",          2560,   note="Main job"); r+=1
    row(r, "Side work  ($15 × 8hr × 4d)", 480,    note="Variable — could flex up or down"); r+=1
    row(r, "Shane (rent + utility share)", 628,    note="⚠ AT RISK — see contingency below"); r+=1
    inc_row = r
    row(r, "TOTAL SPENDABLE", val=INCOME_TOTAL, bold=True, bg=ABLUE); r+=1
    row(r, "Motus reimbursement  (~$372/mo)", note="NOT income — offsets commute gas; don't spend as fun money."); r+=1
    ws.row_dimensions[r].height = 8; r+=1

    sec(r, "ACCOUNT 1 · FIXED  — auto-pay, set & forget", BLUE); r+=1
    row(r, "Mortgage",                 837, bg=LGRAY); r+=1
    row(r, "House insurance",          190); r+=1
    row(r, "Auto insurance",           79,  bg=LGRAY); r+=1
    row(r, "Property tax  (set-aside)",54); r+=1
    row(r, "Phone",                    55,  bg=LGRAY); r+=1
    fixed_r = r
    row(r, "FIXED TOTAL", val=FIXED_TOTAL, bold=True, bg=ABLUE); r+=1
    ws.row_dimensions[r].height = 8; r+=1

    sec(r, "ACCOUNT 2 · FLEX  — fund $630 flat every month", AMBER); r+=1
    row(r, "Water  (avg + 1 SD, 12-mo data)",     90,  bg=LGRAY,
        note="~$45 is a leak hedge — if leak confirmed fixed, drop to $45 and free $45/mo"); r+=1
    row(r, "Power  (avg + 1 SD — peaks Jan/Jul)", 453,
        note="Strong seasonal pattern; flat funding + buffer absorbs swings"); r+=1
    row(r, "Internet  (hedge for promo expiry)",   64,  bg=LGRAY,
        note="Currently $24.95/mo promo — ~$39/mo sweeps to savings until rate changes"); r+=1
    row(r, "Trash",                                23); r+=1
    util_r = r
    row(r, "UTILITIES SUBTOTAL  (fund flat)", val=UTIL_SUB, bold=True, bg=ABLUE); r+=1
    row(r, "Gas  (commute, est.)",   430, bg=LGRAY, note="Motus ~$372 offsets; real net gas ≈ $58"); r+=1
    row(r, "Groceries  (est.)",      400); r+=1
    flex_r = r
    row(r, "FLEX TOTAL", val=FLEX_TOTAL_PY, bold=True, bg=ABLUE); r+=1
    row(r, "⚓  Standing buffer  (never sweep)", 167.63, bg=LGRAY,
        note="Absorbs worst single month across all utilities simultaneously"); r+=1
    ws.row_dimensions[r].height = 8; r+=1

    sec(r, "LEFTOVER  →  splits to savings + spending", GREEN); r+=1
    left_r = r
    row(r, "Leftover after fixed + flex",
        val=LEFTOVER_PY, bold=True, bg=ABLUE); r+=1
    row(r, "⚠  If Shane's $628 stops → leftover",
        val=SHANE_CONT, bold=True, bg="FEF9C3",
        note="PLAN AROUND THIS NUMBER  ·  treat Shane's $628 as a savings bonus if it comes"); r+=1
    ws.row_dimensions[r].height = 8; r+=1

    sec(r, "ACCOUNTS 3 + 4 · SPENDING + SAVINGS", DARK); r+=1
    row(r, "SPENDING — AOD card  (monthly load)", 500, bg=LGRAY,
        note="Only card in your wallet. Balance = your tracker. Lower this to save more."); r+=1
    spend_r = r-1
    row(r, "SAVINGS — SoFi Vaults  (what's left)",
        val=SAVINGS_PY, bg=LGREEN,
        note="Emergency Fund + Move/Career Fund  ·  paid FIRST on payday"); r+=1
    ws.row_dimensions[r].height = 8; r+=1

    sec(r, "BUILT-IN PADDING (sweeps to savings automatically)", "16A34A"); r+=1
    row(r, "Water leak hedge  (if leak fixed, reclaim)",  45,  bg=LGRAY,
        note="Run: python analyze.py exclude Water 2026-04 / 2026-06 to recalculate"); r+=1
    row(r, "Internet promo surplus  (~$39/mo right now)", 39,
        note="Sweeps automatically until rate changes"); r+=1
    row(r, "Hidden savings in flex (approx.)", val=PAD_TOTAL,
        bold=True, bg=LGREEN); r+=1
    ws.row_dimensions[r].height = 8; r+=1

    # ── Self-employment tax reserve ────────────────────────────────
    sec(r, "SELF-EMPLOYMENT TAX RESERVE  (1099 side work)", "7C3AED"); r+=1
    # Row 6 = side work ($480) — known from build order above
    row(r, "Side work gross  (from income above)", val=_SIDE, bg=LGRAY,
        note="1099 — no taxes withheld; this must come out before you touch the money"); r+=1
    se_show_r = r - 1
    se_reserve_r = r
    row(r, "Monthly reserve  (28% — SE tax 15.3% + income tax est.)",
        val=SE_RESERVE_PY, bold=True, bg="EDE9FE",
        note="Transfer on every payday before spending any side income"); r+=1
    row(r, "True spendable from side work",
        val=SE_NET_PY, bg=LGRAY,
        note="What's actually yours after the tax reserve is parked"); r+=1
    ws.row_dimensions[r].height = 8; r+=1

    # ── Sinking funds ─────────────────────────────────────────────
    sec(r, "SINKING FUNDS  — annual / irregular costs ÷ 12", AMBER); r+=1
    row(r, "Truck maintenance  (1999 Silverado, self-maintained)", 75, bg=LGRAY,
        note="~$900/yr est. · adjust to your history"); r+=1
    row(r, "Vehicle registration  (AL, annual)", 17,
        note="~$200/yr ÷ 12"); r+=1
    row(r, "Annual insurance premiums  (if any lump-sum)", 0, bg=LGRAY,
        note="$0 if already paying monthly — update if any biller charges annually"); r+=1
    sink_r = r
    row(r, "SINKING TOTAL  — set aside monthly",
        val=SINK_TOTAL_PY, bold=True, bg=ABLUE,
        note="Park in a labeled savings vault; pull from it when the bill hits"); r+=1
    ws.row_dimensions[r].height = 8; r+=1

    # ── Income trajectory ──────────────────────────────────────────
    sec(r, "INCOME TRAJECTORY  — current effective wage → target", BLUE); r+=1
    row(r, "Current effective wage  (50hr wk + 7hr commute)", bg=LGRAY,
        note="$2,560 take-home ÷ ~217 hrs/mo (50hr work + 7hr commute = ~7hr/day)"); r+=1
    ws[f"C{r-1}"] = 11.81
    ws[f"C{r-1}"].number_format = '"$"#,##0.00'
    ws[f"C{r-1}"].fill          = fill(LGRAY)
    ws[f"C{r-1}"].alignment     = align("right")
    current_th_r = r
    row(r, "Current monthly take-home  (Coke only)", 2560,
        note="Baseline — side work net is on top of this"); r+=1
    current_th_r = r - 1
    row(r, "Target wage after retraining  ($20–25/hr, $22 midpoint)", bg=LGRAY,
        note="Midpoint of stated $20–25/hr range"); r+=1
    ws[f"C{r-1}"] = 22
    ws[f"C{r-1}"].number_format = '"$"#,##0.00'
    ws[f"C{r-1}"].fill          = fill(LGRAY)
    ws[f"C{r-1}"].alignment     = align("right")
    row(r, "Est. take-home at $22/hr  (40hr/wk, ~72% net)",
        val=EST_TH_PY,
        note="Rough estimate — actual varies by deductions"); r+=1
    est_r = r - 1
    row(r, "Monthly income gain at target",
        val=INCOME_GAIN_PY, bold=True, bg=LGREEN,
        note="Extra per month once you land the better role"); r+=1
    ws.row_dimensions[r].height = 8; r+=1

    # ── Savings goals ──────────────────────────────────────────────
    sec(r, "SAVINGS GOALS  (sequential — ① Emergency Fund first)", GREEN); r+=1
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"A{r}:B{r}")
    ws[f"A{r}"] = "Current savings balance  ← UPDATE THIS INPUT"
    ws[f"A{r}"].fill      = fill(LBLUE)
    ws[f"A{r}"].font      = Font(bold=True, color=BLUE, size=10, name="Calibri")
    ws[f"A{r}"].alignment = align("left", "center")
    ws[f"C{r}"] = 502.15   # savings $250 + vault $252.15 (update when balance changes)
    ws[f"C{r}"].number_format = '"$"#,##0.00'
    ws[f"C{r}"].font          = Font(bold=True, color=BLUE, size=10, name="Calibri")
    ws[f"C{r}"].fill          = fill(LBLUE)
    ws[f"C{r}"].alignment     = align("right", "center")
    ws[f"D{r}"] = "Blue = input cell  ·  enter your actual savings balance here"
    ws[f"D{r}"].font      = Font(italic=True, size=9, color=BLUE, name="Calibri")
    ws[f"D{r}"].fill      = fill(LBLUE)
    ws[f"D{r}"].alignment = align(wrap=True)
    savings_r = r; r+=1
    ws.row_dimensions[r].height = 6; r+=1

    row(r, "① EMERGENCY FUND — target", 7000, bold=True, bg=LGRAY); r+=1
    ef_tgt_r = r - 1
    row(r, "   Current  (from balance above)",
        fx=f"=MIN(C{savings_r},7000)", bg=LGRAY); r+=1
    ef_cur_r = r - 1
    row(r, "   Remaining", fx=f"=MAX(0,C{ef_tgt_r}-C{ef_cur_r})", bg=LGRAY); r+=1
    row(r, "   % Complete",
        fx=f"=IFERROR(C{ef_cur_r}/C{ef_tgt_r},0)", bold=True, bg=LGREEN); r+=1
    ws[f"C{r-1}"].number_format = "0.0%"
    ws.row_dimensions[r].height = 6; r+=1

    row(r, "② SEPTIC FUND — target  (starts after EF hits $7,000)", 7285, bold=True); r+=1
    sep_tgt_r = r - 1
    row(r, "   Current  (balance above − $7,000 EF)",
        fx=f"=MAX(0,C{savings_r}-7000)"); r+=1
    sep_cur_r = r - 1
    row(r, "   Remaining", fx=f"=MAX(0,C{sep_tgt_r}-C{sep_cur_r})", bg=LGRAY); r+=1
    row(r, "   % Complete",
        fx=f"=IFERROR(C{sep_cur_r}/C{sep_tgt_r},0)", bold=True, bg=LGREEN); r+=1
    ws[f"C{r-1}"].number_format = "0.0%"
    ws.row_dimensions[r].height = 8; r+=1

    # ── Revised bottom line (accounts for everything) ──────────────
    sec(r, "REVISED MONTHLY PICTURE  (after tax reserve + sinking funds)", DARK); r+=1
    row(r, "Gross leftover  (after fixed + flex)", val=LEFTOVER_PY, bg=LGRAY); r+=1
    row(r, "  Less: SE tax reserve", val=-SE_RESERVE_PY); r+=1
    row(r, "  Less: sinking funds", val=-SINK_TOTAL_PY, bg=LGRAY); r+=1
    true_left_r = r
    row(r, "TRUE MONTHLY AVAILABLE",
        val=TRUE_AVAIL_PY, bold=True, bg=ABLUE); r+=1
    row(r, "  If Shane's $628 stops → true available",
        val=SHANE_FL_PY, bold=True, bg="FEF9C3",
        note="Hard number — build the system around this floor"); r+=1

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Monthly Check-In
# ══════════════════════════════════════════════════════════════════════════════
def build_checkin(wb, monthly):
    ws = wb.create_sheet("Monthly Check-In")
    ws.sheet_view.showGridLines = False

    col_w = {"A":11,"B":10,"C":10,"D":11,"E":9,
              "F":13,"G":10,"H":14,"I":14,"J":13,"K":22}
    for c, w in col_w.items():
        ws.column_dimensions[c].width = w

    # ── Title ─────────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    ws["A1"] = "MONTHLY CHECK-IN & SWEEP LOG"
    ws["A1"].fill      = fill(DARK)
    ws["A1"].font      = Font(bold=True, size=16, color=WHITE, name="Calibri")
    ws["A1"].alignment = align("left","center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:K2")
    ws["A2"] = (f"Fund $630 flat every month · fill in actual bills · "
                f"green = surplus sweep to savings · red = buffer drawn")
    ws["A2"].fill      = fill(LGRAY)
    ws["A2"].font      = font(italic=True, size=9, color=SUB)
    ws["A2"].alignment = align("left","center")
    ws.row_dimensions[2].height = 18

    # ── Column headers ─────────────────────────────────────────────
    HDR = 4
    labels = ["Month","Water","Power","Internet","Trash",
              "Total Actual","Funded","Net  (+/−)","Buffer Bal","→ Savings","Notes"]
    colors = [DARK, UTIL_HEX["Water"], UTIL_HEX["Power"], UTIL_HEX["Internet"],
              UTIL_HEX["Trash"], DARK, BLUE, "334155", AMBER, GREEN, DARK]
    header_row(ws, HDR, list(range(1,12)), labels, colors, height=28)

    # ── Pre-compute historical row values ──────────────────────────
    # Run the buffer/sweep math in Python for every month that has real data
    # so historical rows display correctly in any viewer (not just Excel).
    running_buf = BUFFER
    precomp = {}  # date_str → {total, net, buf, sweep}
    for date_str in sorted(monthly.keys()):
        data  = monthly[date_str]
        total = round(sum((data.get(u) or 0) for u in UTILITIES), 2)
        net   = round(FUNDED - total, 2)
        new_buf = round(max(0, min(BUFFER, running_buf + net)), 2)
        refill  = max(0, BUFFER - running_buf)
        sweep   = round(max(0, net - refill), 2)
        precomp[date_str] = {"total": total, "net": net, "buf": new_buf, "sweep": sweep}
        running_buf = new_buf

    # ── Build month list ────────────────────────────────────────────
    D = HDR + 1   # first data row
    all_months = list(monthly.keys())

    # Insert May 2026 placeholder if missing
    if "2026-05" not in all_months:
        idx = all_months.index("2026-04") + 1 if "2026-04" in all_months else len(all_months)
        all_months.insert(idx, "2026-05")

    last_hist  = all_months[-1] if all_months else "2026-06"
    extra      = future_months(last_hist, n=18)
    all_rows   = all_months + extra
    hist_count = len(all_months)

    prev_buf_row    = None    # Excel row index of most-recent buffer cell (for formulas)
    prev_actual_buf = BUFFER  # Python-tracked buffer for blank/carry-forward rows

    for i, date_str in enumerate(all_rows):
        r        = D + i
        pc       = precomp.get(date_str)      # pre-computed values; None if no data
        is_blank = (date_str == "2026-05")    # known missing month placeholder
        bg       = WHITE if i % 2 == 0 else LGRAY
        ws.row_dimensions[r].height = 20

        data = monthly.get(date_str, {u: None for u in UTILITIES})

        # A: Month label
        cell(ws, f"A{r}", mlabel(date_str), bold=True, size=9, bg=bg, ha="center")

        # B-E: Utility actuals (pre-filled for historical, blank for future)
        for j, util in enumerate(UTILITIES, 2):
            col = get_column_letter(j)
            val = data.get(util) if date_str in monthly else None
            v   = val if (val is not None and val > 0) else None
            ws[f"{col}{r}"].value         = v
            ws[f"{col}{r}"].fill          = fill(bg)
            ws[f"{col}{r}"].font          = font(size=9)
            ws[f"{col}{r}"].alignment     = align("center","center")
            ws[f"{col}{r}"].number_format = '"$"#,##0.00'

        # F: Total Actual
        if pc:
            ws[f"F{r}"].value = pc["total"]
        else:
            # IF(COUNTA(...)=0,"",SUM(...)) so blank months don't show $0
            ws[f"F{r}"] = f'=IF(COUNTA(B{r}:E{r})=0,"",SUM(B{r}:E{r}))'
        ws[f"F{r}"].fill          = fill(bg)
        ws[f"F{r}"].font          = Font(bold=True, size=9, color=TEXT, name="Calibri")
        ws[f"F{r}"].alignment     = align("center","center")
        ws[f"F{r}"].number_format = '"$"#,##0.00'

        # G: Funded flat
        ws[f"G{r}"] = FUNDED
        ws[f"G{r}"].fill          = fill(bg)
        ws[f"G{r}"].font          = Font(size=9, color=BLUE, name="Calibri")
        ws[f"G{r}"].alignment     = align("center","center")
        ws[f"G{r}"].number_format = '"$"#,##0.00'

        # H: Net = Funded − Actual
        if pc:
            ws[f"H{r}"].value = pc["net"]
        else:
            ws[f"H{r}"] = f'=IF(F{r}="","",G{r}-F{r})'
        ws[f"H{r}"].fill          = fill(bg)
        ws[f"H{r}"].font          = Font(bold=True, size=9, color=TEXT, name="Calibri")
        ws[f"H{r}"].alignment     = align("center","center")
        ws[f"H{r}"].number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'

        # I: Buffer Balance  MAX(0, MIN(cap, prev_buf + net))
        # J: Sweep → Savings  MAX(0, net − MAX(0, cap − prev_buf))
        # Both computed BEFORE updating prev_buf_row so J can reference the
        # correct previous-month buffer (fixing the original sweep formula bug).
        if pc:
            ws[f"I{r}"].value = pc["buf"]
            ws[f"J{r}"].value = pc["sweep"]
            prev_actual_buf   = pc["buf"]
        elif is_blank:
            ws[f"I{r}"].value = prev_actual_buf  # carry forward, no bills this month
            ws[f"J{r}"].value = None
        else:
            # Future row — formulas reference previous buffer row (prev_buf_row),
            # NOT the current row, so refill math is correct.
            if prev_buf_row is None:
                buf_fx   = f'=IF(H{r}="","",MAX(0,MIN({BUFFER},{BUFFER}+H{r})))'
                sweep_fx = f'=IF(H{r}="","",MAX(0,H{r}))'
            else:
                buf_fx   = f'=IF(H{r}="","",MAX(0,MIN({BUFFER},I{prev_buf_row}+H{r})))'
                sweep_fx = (f'=IF(H{r}="","",MAX(0,H{r}'
                            f'-MAX(0,{BUFFER}-I{prev_buf_row})))')
            ws[f"I{r}"] = buf_fx
            ws[f"J{r}"] = sweep_fx

        for col in ["I", "J"]:
            ws[f"{col}{r}"].fill      = fill(bg)
            ws[f"{col}{r}"].alignment = align("center","center")
        ws[f"I{r}"].font          = font(size=9)
        ws[f"I{r}"].number_format = '"$"#,##0.00'
        ws[f"J{r}"].font          = Font(size=9, color=GREEN, name="Calibri")
        ws[f"J{r}"].number_format = '"$"#,##0.00'

        # K: Notes
        ws[f"K{r}"].fill      = fill(bg)
        ws[f"K{r}"].font      = font(size=9, color=SUB, italic=True)
        ws[f"K{r}"].alignment = align("left","center")

        prev_buf_row = r  # update AFTER I and J are written

    # ── Totals row ─────────────────────────────────────────────────
    last_r   = D + len(all_rows) - 1
    hist_end = D + hist_count - 1
    TOT = last_r + 2
    ws.row_dimensions[TOT].height = 24
    ws[f"A{TOT}"] = "TOTALS"
    ws[f"A{TOT}"].fill      = fill(DARK)
    ws[f"A{TOT}"].font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
    ws[f"A{TOT}"].alignment = align("center","center")

    # Pre-compute sums from the precomp dict (excludes May '26 placeholder)
    hist_months = sorted(monthly.keys())
    col_sums = {
        "B": sum(monthly[m].get("Water")    or 0 for m in hist_months),
        "C": sum(monthly[m].get("Power")    or 0 for m in hist_months),
        "D": sum(monthly[m].get("Internet") or 0 for m in hist_months),
        "E": sum(monthly[m].get("Trash")    or 0 for m in hist_months),
        "F": sum(precomp[m]["total"] for m in hist_months),
        "H": sum(precomp[m]["net"]   for m in hist_months),
        "J": sum(precomp[m]["sweep"] for m in hist_months),
    }
    for c, v in col_sums.items():
        ws[f"{c}{TOT}"].value         = round(v, 2)
        ws[f"{c}{TOT}"].fill          = fill(DARK)
        ws[f"{c}{TOT}"].font          = Font(bold=True, color=WHITE, size=10, name="Calibri")
        ws[f"{c}{TOT}"].alignment     = align("center","center")
        ws[f"{c}{TOT}"].number_format = '"$"#,##0.00'
    ws[f"G{TOT}"].value         = round(FUNDED * len(hist_months), 2)
    ws[f"G{TOT}"].fill          = fill(DARK)
    ws[f"G{TOT}"].font          = Font(bold=True, color=WHITE, size=10, name="Calibri")
    ws[f"G{TOT}"].alignment     = align("center","center")
    ws[f"G{TOT}"].number_format = '"$"#,##0.00'

    # ── Averages row ────────────────────────────────────────────────
    AVG = TOT + 1
    n   = len(hist_months)
    ws.row_dimensions[AVG].height = 20
    ws[f"A{AVG}"] = "AVERAGES"
    ws[f"A{AVG}"].fill      = fill(LGRAY)
    ws[f"A{AVG}"].font      = Font(bold=True, size=9, color=TEXT, name="Calibri")
    ws[f"A{AVG}"].alignment = align("center","center")
    avg_vals = {
        "B": round(col_sums["B"] / n, 2),
        "C": round(col_sums["C"] / n, 2),
        "D": round(col_sums["D"] / n, 2),
        "E": round(col_sums["E"] / n, 2),
        "F": round(col_sums["F"] / n, 2),
        "G": FUNDED,
        "H": round(col_sums["H"] / n, 2),
        "J": round(col_sums["J"] / n, 2),
    }
    for c, v in avg_vals.items():
        ws[f"{c}{AVG}"].value         = v
        ws[f"{c}{AVG}"].fill          = fill(LGRAY)
        ws[f"{c}{AVG}"].font          = font(size=9)
        ws[f"{c}{AVG}"].alignment     = align("center","center")
        ws[f"{c}{AVG}"].number_format = '"$"#,##0.00'

    # ── Conditional formatting on Net column ───────────────────────
    gf = PatternFill("solid", fgColor=LGREEN)
    rf = PatternFill("solid", fgColor=LRED)
    gn = Font(color="065F46", bold=True, size=9, name="Calibri")
    rn = Font(color="991B1B", bold=True, size=9, name="Calibri")
    ws.conditional_formatting.add(
        f"H{D}:H{hist_end}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=gf, font=gn))
    ws.conditional_formatting.add(
        f"H{D}:H{hist_end}",
        CellIsRule(operator="lessThan", formula=["0"], fill=rf, font=rn))
    # Also on Sweep column
    ws.conditional_formatting.add(
        f"J{D}:J{hist_end}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=gf, font=gn))

    # ── Embed charts ───────────────────────────────────────────────
    img_row_anchor = AVG + 3
    charts = [
        ("total.png",    f"A{img_row_anchor}",      720, 330),
        ("overview.png", f"A{img_row_anchor + 22}",  720, 400),
        ("power.png",    f"A{img_row_anchor + 49}",  680, 300),
        ("water.png",    f"H{img_row_anchor + 49}",  680, 300),
        ("internet.png", f"A{img_row_anchor + 69}",  680, 300),
        ("trash.png",    f"H{img_row_anchor + 69}",  680, 300),
    ]
    for fname, anchor, w, h in charts:
        embed_image(ws, os.path.join(CHARTS_DIR, fname), anchor, w, h)

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Per Utility Detail
# ══════════════════════════════════════════════════════════════════════════════
def build_per_utility(wb, monthly):
    ws = wb.create_sheet("Per Utility")
    ws.sheet_view.showGridLines = False

    for c, w in {"A":12,"B":10,"C":10,"D":10,"E":14,"F":22}.items():
        ws.column_dimensions[c].width = w

    ws.merge_cells("A1:F1")
    ws["A1"] = "PER UTILITY — Actuals vs Targets"
    ws["A1"].fill      = fill(DARK)
    ws["A1"].font      = Font(bold=True, size=14, color=WHITE, name="Calibri")
    ws["A1"].alignment = align("left","center")
    ws.row_dimensions[1].height = 32

    targets = {"Water": 90, "Power": 453, "Internet": 64, "Trash": 23}
    avgs    = {"Water": 49, "Power": 344, "Internet": 42, "Trash": 22}

    start = 3
    for util in UTILITIES:
        clr    = UTIL_HEX[util]
        target = targets[util]
        avg    = avgs[util]

        # Section header
        ws.row_dimensions[start].height = 26
        ws.merge_cells(f"A{start}:F{start}")
        ws[f"A{start}"] = f"{util.upper()}  ·  fund ${target}/mo  ·  avg ${avg}  ·  cushion ${target-avg}/mo"
        ws[f"A{start}"].fill      = fill(clr)
        ws[f"A{start}"].font      = Font(bold=True, size=11, color=WHITE, name="Calibri")
        ws[f"A{start}"].alignment = align("left","center")

        # Column headers
        h = start + 1
        ws.row_dimensions[h].height = 20
        for col, lbl in zip(["A","B","C","D","E"],
                             ["Month","Actual","Target","Net","Notes"]):
            ws[f"{col}{h}"] = lbl
            ws[f"{col}{h}"].fill      = fill(LGRAY)
            ws[f"{col}{h}"].font      = Font(bold=True, size=9, color=TEXT, name="Calibri")
            ws[f"{col}{h}"].alignment = align("center","center")

        # Data
        d_start = h + 1
        months  = sorted(monthly.keys())
        for i, date_str in enumerate(months):
            r  = d_start + i
            bg = WHITE if i % 2 == 0 else LGRAY
            val = monthly[date_str].get(util)
            ws.row_dimensions[r].height = 18

            cell(ws, f"A{r}", mlabel(date_str), size=9, bg=bg, ha="center")

            ws[f"B{r}"].value         = val if (val and val > 0) else None
            ws[f"B{r}"].fill          = fill(bg)
            ws[f"B{r}"].font          = font(size=9)
            ws[f"B{r}"].alignment     = align("center","center")
            ws[f"B{r}"].number_format = '"$"#,##0.00'

            ws[f"C{r}"] = target
            ws[f"C{r}"].fill          = fill(bg)
            ws[f"C{r}"].font          = Font(size=9, color=SUB, name="Calibri")
            ws[f"C{r}"].alignment     = align("center","center")
            ws[f"C{r}"].number_format = '"$"#,##0.00'

            ws[f"D{r}"] = f'=IFERROR(C{r}-B{r},"")'
            ws[f"D{r}"].fill          = fill(bg)
            ws[f"D{r}"].font          = font(size=9)
            ws[f"D{r}"].alignment     = align("center","center")
            ws[f"D{r}"].number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'

            ws[f"E{r}"].fill      = fill(bg)
            ws[f"E{r}"].font      = font(size=9, color=SUB, italic=True)
            ws[f"E{r}"].alignment = align("left","center")

        # Totals
        tot = d_start + len(months)
        ws.row_dimensions[tot].height = 22
        for c in ["A","B","C","D"]:
            ws[f"{c}{tot}"].fill      = fill(LGRAY)
            ws[f"{c}{tot}"].font      = Font(bold=True, size=9, name="Calibri")
            ws[f"{c}{tot}"].alignment = align("center","center")
        ws[f"A{tot}"] = "TOTALS"
        ws[f"B{tot}"] = f"=SUM(B{d_start}:B{d_start+len(months)-1})"
        ws[f"B{tot}"].number_format = '"$"#,##0.00'
        ws[f"C{tot}"] = target * len(months)
        ws[f"C{tot}"].number_format = '"$"#,##0.00'
        ws[f"D{tot}"] = f"=C{tot}-B{tot}"
        ws[f"D{tot}"].number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'

        # Conditional on Net
        gf = PatternFill("solid", fgColor=LGREEN)
        rf = PatternFill("solid", fgColor=LRED)
        gn = Font(color="065F46", bold=True, size=9, name="Calibri")
        rn = Font(color="991B1B", bold=True, size=9, name="Calibri")
        ws.conditional_formatting.add(
            f"D{d_start}:D{d_start+len(months)-1}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=gf, font=gn))
        ws.conditional_formatting.add(
            f"D{d_start}:D{d_start+len(months)-1}",
            CellIsRule(operator="lessThan",    formula=["0"], fill=rf, font=rn))

        start = tot + 3

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Bill Calendar
# ══════════════════════════════════════════════════════════════════════════════
def build_calendar(wb):
    ws = wb.create_sheet("Bill Calendar")
    ws.sheet_view.showGridLines = False
    for c, w in {"A":2,"B":28,"C":12,"D":15,"E":15,"F":30}.items():
        ws.column_dimensions[c].width = w

    ws.merge_cells("A1:F1")
    ws["A1"] = "BILL CALENDAR — ALIGN TO PAYDAY"
    ws["A1"].fill      = fill(DARK)
    ws["A1"].font      = Font(bold=True, size=14, color=WHITE, name="Calibri")
    ws["A1"].alignment = align("left","center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"] = "Move due dates so bills land 2–3 days after payday · keep buffers · set it once"
    ws["A2"].fill      = fill(LGRAY)
    ws["A2"].font      = font(italic=True, size=9, color=SUB)
    ws["A2"].alignment = align("left","center")

    r = 4
    for col, h in zip(["B","C","D","E","F"],
                       ["BILL","AMOUNT","CURRENT DUE","TARGET DUE","NOTES"]):
        ws[f"{col}{r}"] = h
        ws[f"{col}{r}"].fill      = fill(DARK)
        ws[f"{col}{r}"].font      = Font(bold=True, color=WHITE, size=9, name="Calibri")
        ws[f"{col}{r}"].alignment = align("center","center")
    ws.row_dimensions[r].height = 22; r += 1

    def bill_section(title, bg, bills):
        nonlocal r
        ws.row_dimensions[r].height = 22
        ws.merge_cells(f"B{r}:F{r}")
        ws[f"B{r}"] = title
        ws[f"B{r}"].fill      = fill(bg)
        ws[f"B{r}"].font      = Font(bold=True, color=WHITE, size=9, name="Calibri")
        ws[f"B{r}"].alignment = align("left","center")
        r += 1
        for i, (name, amt, note) in enumerate(bills):
            bg2 = LGRAY if i % 2 == 0 else WHITE
            ws.row_dimensions[r].height = 20
            for c in ["B","C","D","E","F"]:
                ws[f"{c}{r}"].fill = fill(bg2)
            ws[f"B{r}"] = name
            ws[f"C{r}"] = amt
            ws[f"D{r}"] = "—"
            ws[f"E{r}"] = "—"
            ws[f"F{r}"] = note
            ws[f"B{r}"].font      = font(size=9)
            ws[f"C{r}"].font      = font(size=9)
            ws[f"C{r}"].number_format = '"$"#,##0.00'
            ws[f"D{r}"].font      = font(size=9, color=SUB)
            ws[f"E{r}"].font      = font(size=9, color=BLUE)
            ws[f"F{r}"].font      = font(size=9, color=SUB, italic=True)
            for c in ["B","C","D","E","F"]:
                ws[f"{c}{r}"].alignment = align("center","center")
            ws[f"B{r}"].alignment = align()
            ws[f"F{r}"].alignment = align(wrap=True)
            r += 1
        r += 1

    bill_section("FIXED BILLS — auto-pay from SoFi checking", BLUE, [
        ("Mortgage",               837, "Move to 3rd of month"),
        ("House insurance",        190, "Move to 4th"),
        ("Auto insurance",         79,  "Move to 5th"),
        ("Phone",                  55,  "Move to 6th"),
        ("Property tax set-aside", 54,  "Move to 7th — manual transfer"),
    ])
    bill_section("FLEX UTILITIES — auto-draft from Flex account", AMBER, [
        ("Power",    453, "Varies — flat funding absorbs swings"),
        ("Water",    90,  "Flat funded; two outlier months flagged"),
        ("Internet", 64,  "On promo $24.95 — $39/mo currently sweeps"),
        ("Trash",    23,  "Stable — rarely varies"),
    ])

    # How-to notes
    r += 1
    ws.merge_cells(f"B{r}:F{r}")
    ws[f"B{r}"] = "HOW TO ALIGN DUE DATES — 3 moves"
    ws[f"B{r}"].fill      = fill(DARK)
    ws[f"B{r}"].font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
    ws[f"B{r}"].alignment = align()
    ws.row_dimensions[r].height = 22; r += 1

    steps = [
        ("Move the due dates",
         "Call each biller and ask to shift your due date — almost all say yes. "
         "Target 2-3 days after payday so money is always there when the bill hits."),
        ("Keep buffers",
         f"~$1,215 (one month fixed) parked in SoFi checking + "
         f"${BUFFER:.0f} in the Flex account. Once those cushions exist, timing stops mattering."),
        ("Group by paycheck",
         "If paid biweekly, assign bills to specific checks so no single check is over-loaded. "
         "Mortgage from the check that clears it; utilities auto-draft mid-month."),
    ]
    for i, (title, body) in enumerate(steps):
        bg = LGRAY if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 15
        ws[f"B{r}"] = f"● {title}"
        ws[f"B{r}"].fill      = fill(bg)
        ws[f"B{r}"].font      = Font(bold=True, size=9, name="Calibri")
        ws[f"B{r}"].alignment = align()
        r += 1
        ws.merge_cells(f"B{r}:F{r}")
        ws[f"B{r}"] = body
        ws[f"B{r}"].fill      = fill(bg)
        ws[f"B{r}"].font      = font(size=9, color=SUB, italic=True)
        ws[f"B{r}"].alignment = align(wrap=True)
        ws.row_dimensions[r].height = 36
        r += 1

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    rows    = load_bills()
    monthly = get_monthly(rows)

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    build_system(wb)
    build_checkin(wb, monthly)
    build_per_utility(wb, monthly)
    build_calendar(wb)

    wb.save(OUT_FILE)
    print(f"Saved → {OUT_FILE}")

if __name__ == "__main__":
    main()
